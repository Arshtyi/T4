import notices_check
import urllib.request
from urllib.parse import urljoin
from bs4 import BeautifulSoup
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
import schedule
import time
import subprocess
import apscheduler
from apscheduler.schedulers.blocking import BlockingScheduler
from datetime import datetime

if __name__ == '__main__':
    
    #创建工作簿,子表
    workbook = op.Workbook()
    worksheet = workbook.active
    #格式设置
    #字体
    ft = Font(name=u'微软雅黑',
        size=11,
        bold=True,
        strike=False ,#删除线
        )
        
    #对齐方式
    alignment=Alignment(horizontal = 'center',#水平
            vertical = 'center',#垂直
            text_rotation=0,#旋转角度0~180
            wrap_text=True,#文字换行
            shrink_to_fit=True,#自适应宽度,改变文字大小,上一项false
            indent=0)
        
    number_format = 'General'
        
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 60
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 100
    worksheet.column_dimensions['E'].width = 120#####调整列宽
    worksheet.append( ['通知来源','通知链接','通知发布时间','通知标题','通知简要'])##表头
    workbook.save('./ret.xlsx')
    row_n = 2##记录行号

    while True:
        print("选择你要爬取的网站")
        print("1 : 山大视点-山大要闻;")
        print("2 : 本科生院-工作通知;")
        print("3 : 计算机学院-本科教育;")
        print("4 : 在下一个自动爬取点自动爬取；")
        print("5 : 停止;")
        print("6 : 重置ret.xlsx文件.")#功能选项
        choice = (int)(input())
        if choice == 4:#自动爬取
            print("程序将在上午7点,正午12点,下午6点,晚上10点自动运行.")
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"当前时间：{current_time}")
            print("Ctrl+C以停止程序.")
            scheduler = BlockingScheduler()
            try:
                scheduler.add_job(notices_check.Auto_fetch,'cron',hour = 7,minute = 0)
                scheduler.add_job(notices_check.Auto_fetch,'cron',hour = 12,minute =0)
                scheduler.add_job(notices_check.Auto_fetch,'cron',hour = 18,minute = 0)
                scheduler.add_job(notices_check.Auto_fetch,'cron',hour = 22,minute = 0)
                scheduler.start()
            except KeyboardInterrupt:#停止信号
                print("程序已终止.")
        elif choice == 5:
            print("已停止.")
            workbook = load_workbook(filename = r'./ret.xlsx')
            worksheet = workbook.active
            for row in worksheet.iter_rows(min_row=1,max_row=1000,min_col=1,max_col=50):
                for cell in row:
                    cell.font = ft
                    cell.alignment = alignment
                    cell.number_format = number_format
            workbook.save('./ret.xlsx')
            break#格式化后停止
        else  :
            if choice == 6:#重置操作
                workbook = load_workbook(filename = r'./ret.xlsx')
                worksheet = workbook.active
                del workbook['Sheet']
                workbook = op.Workbook()
                worksheet = workbook.active
                worksheet.column_dimensions['A'].width = 20
                worksheet.column_dimensions['B'].width = 60
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 100
                worksheet.column_dimensions['E'].width = 120#####调整列宽
                worksheet.append( ['通知来源','通知链接','通知发布时间','通知标题','通知简要'])##表头
                row_n = 2##记录行号
                for row in worksheet.iter_rows(min_row=1,max_row=1000,min_col=1,max_col=50):
                    for cell in row:
                        cell.font = ft
                        cell.alignment = alignment
                        cell.number_format = number_format
                workbook.save('ret.xlsx')#格式化
            else:#爬取操作
                numbers = (int)(input("你期望的数量（以页计）："))
                workbook = load_workbook(filename = r'./ret.xlsx')
                worksheet = workbook.active
                row_n = worksheet.max_row + 1
                workbook.save('ret.xlsx')
                notices_check.fetch(numbers,choice,row_n)
