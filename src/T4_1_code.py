import urllib.request
from urllib.parse import urljoin
from bs4 import BeautifulSoup
import openpyxl as op
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
import schedule
import time
import subprocess
import apscheduler
from apscheduler.schedulers.blocking import BlockingScheduler
from datetime import datetime
def fetch(numbers,choice,row_n):
    workbook = load_workbook(filename = r'./ret.xlsx')
    worksheet = workbook.active#取出表
    st_row_n = row_n#找到操作行
    if choice == 1:#山大视点-山大要闻
        start = 692 - numbers#依据页数做起始处理
        for i in range(start,692):
            i = start + 691 - i#倒序
            if(i==691):#翻页处理
                url = 'https://www.view.sdu.edu.cn/zhxw.htm'
            else:
                url = 'https://www.view.sdu.edu.cn/zhxw/'+str(i)+'.htm'#根据页码对url处理
            # 读取给定 url 的 html 代码
            response = urllib.request.urlopen(url)
            content = response.read().decode('utf-8')
            # 转换读取到的 html 文档
            soup = BeautifulSoup(content, 'html.parser', from_encoding='utf-8')
            # 获取转换后的 html 文档里属性 box14 的 div 标签的内容
            divs = soup.find_all('div', {'class': "box14"})
            # 从已获取的 div 标签的内容里获取 li 标签的内容
            lis = divs[0].find_all('li')
            # 遍历获取到的 lis 列表，并从中抓取链接,标题,时间
            for li in lis:
                url1 = "https://www.view.sdu.edu.cn/"
                url2 = li.find_all('a')[0].get("href")
                # 使用urllib的urljoin()拼接两个地址
                # urljoin的第一个参数是基础母站的url, 第二个是需要拼接成绝对路径的url
                # 利用urljoin，我们可以将爬取的url的相对路径拼接成绝对路径
                url = urljoin(url1, url2)
                fetitle = li.find_all('a')[0].get("title")
                date = li.find_all('span')[0].get_text()
                # 爬取的链接，标题，时间
                #_in 内部#
                response_in = urllib.request.urlopen(url)
                content_in = response_in.read().decode('utf-8')
                #通知具体内容#_in#
                soup_in = BeautifulSoup(content_in,'html.parser',from_encoding='utf-8')
                divs_in = soup_in.find_all('div',{'class':"news_content"})
                newscontent = divs_in[0].find_all('p')[0].get_text()
                #取出通知简要
                insertdata = ['山大视点-山大要闻',url,date,fetitle,newscontent]
                #拼接为插入信息
                worksheet.append(insertdata)
                #追加到尾部
                row_n += 1
                #记录行号
        worksheet.merge_cells(start_column = 1,start_row = st_row_n,end_column = 1,end_row = row_n - 1 )
        #合并通知来源项
        
    elif choice == 2:#本科生院-工作通知
        start = 181 - numbers
        for i in range(start,181):
            i = start + 180 - i
            if(i==180):
                url = 'https://www.bkjx.sdu.edu.cn/index/gztz.htm'
            else:
                url = 'https://www.bkjx.sdu.edu.cn/index/gztz/'+str(i)+'.htm'#根据页码对url处理
            # 读取给定 url 的 html 代码
            response = urllib.request.urlopen(url)
            content = response.read().decode('utf-8')
            # 转换读取到的 html 文档
            soup = BeautifulSoup(content, 'html.parser', from_encoding='utf-8')
            # 获取转换后的 html 文档里属性 newscontent 的 div 标签的内容
            divs = soup.find_all('div', {'class': "newscontent"})
            # 分别取得标题和时间的 div 标签
            divs1 = divs[0].find_all('div',{'style':"float:left"})
            divs2 = divs[0].find_all('div',{'style':"float:right;"})
            for div1 in divs1:
                url1 = "https://www.bkjx.sdu.edu.cn/"
                url2 = div1.find_all('a')[0].get("href")
                # 使用urllib的urljoin()拼接两个地址
                # urljoin的第一个参数是基础母站的url, 第二个是需要拼接成绝对路径的url
                # 利用urljoin，我们可以将爬取的url的相对路径拼接成绝对路径
                url = urljoin(url1, url2)
                fetitle = div1.find_all('a')[0].get("title")
                temp = divs1.index(div1)
                date = divs2[temp].get_text()
                # 爬取的链接，标题,时间
                insertdata = ['本科生院-工作通知',url,date,fetitle]
                # #拼接
                worksheet.append(insertdata)
                #追加
                row_n += 1
                #记录行号
        worksheet.merge_cells(start_column = 1,start_row = st_row_n,end_column = 1,end_row = row_n - 1 )
        #合并通知来源项
    
    elif choice == 3:#计算机学院-本科教育
        start = 7 - numbers
        for i in range(start,7):
            i = start + 6 - i
            if(i==6):
                url = 'https://www.cs.sdu.edu.cn/bkjy.htm'
            else:
                url = 'https://www.cs.sdu.edu.cn/bkjy/'+str(i)+'.htm'#根据页码对url处理
            # 读取给定 url 的 html 代码
            response = urllib.request.urlopen(url)
            content = response.read().decode('utf-8')
            #print(content)
            # 转换读取到的 html 文档
            soup = BeautifulSoup(content, 'html.parser', from_encoding='utf-8')
            # 获取转换后的 html 文档里属性 dplb 的 div 标签的内容
            divs = soup.find_all('div', {'class': "dqlb"})
            lis = divs[0].find_all('li')#取得li标签后做遍历
            for li in lis:
                url1 = "https://www.cs.sdu.edu.cn"
                url2 = li.find_all('a')[0].get("href")
                # 使用urllib的urljoin()拼接两个地址
                # urljoin的第一个参数是基础母站的url, 第二个是需要拼接成绝对路径的url
                # 利用urljoin，我们可以将爬取的url的相对路径拼接成绝对路径
                url = urljoin(url1, url2)
                fetitle = li.find_all('a')[0].get("title")
                date = li.find_all('span')[0].get_text()
                # 爬取的链接，标题,时间
                insertdata = ['计算机学院-本科教育',url,date,fetitle]
                # #拼接
                worksheet.append(insertdata)
                #追加
                row_n += 1
                #记录行号
        worksheet.merge_cells(start_column = 1,start_row = st_row_n,end_column = 1,end_row = row_n - 1 )
        #合并通知来源项
    
    workbook.save('./ret.xlsx')#关闭

def Auto_fetch():#自动爬取
    print("fetch is running ...")
    workbook = load_workbook(filename = r'./ret.xlsx')
    worksheet = workbook.active#取表
    row_n = worksheet.max_row + 1#找到操作行
    workbook.save('ret.xlsx')#关闭
    fetch(1,1,row_n)
    print("已爬取\"山大视点-山大要闻\";")#爬取
    workbook = load_workbook(filename = r'./ret.xlsx')
    worksheet = workbook.active
    row_n = worksheet.max_row + 1
    workbook.save('ret.xlsx')
    fetch(1,2,row_n)
    print("已爬取\"本科生院-工作通知\";")
    workbook = load_workbook(filename = r'./ret.xlsx')
    worksheet = workbook.active
    row_n = worksheet.max_row + 1
    workbook.save('ret.xlsx')
    fetch(1,3,row_n)
    print("已爬取\"计算机学院-本科教育\";")
    print("自动爬取结束,程序已停止.")
    workbook = load_workbook(filename = r'./ret.xlsx')
    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=1,max_row=1000,min_col=1,max_col=50):
        for cell in row:
            cell.font = ft
            cell.alignment = alignment
            cell.number_format = number_format
    workbook.save('./ret.xlsx')#格式化


headers = {
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36 Edg/128.0.0.0"
}


#创建工作簿，子表
workbook = op.Workbook()
worksheet = workbook.active
#格式设置
#字体
ft = Font(name=u'微软雅黑',
    size=11,
    bold=True,
    strike=False ,#删除线
    )
    
#对齐方式
alignment=Alignment(horizontal = 'center',#水平
        vertical = 'center',#垂直
        text_rotation=0,#旋转角度0~180
        wrap_text=True,#文字换行
        shrink_to_fit=True,#自适应宽度，改变文字大小,上一项false
        indent=0)
    
number_format = 'General'
    
worksheet.column_dimensions['A'].width = 20
worksheet.column_dimensions['B'].width = 60
worksheet.column_dimensions['C'].width = 15
worksheet.column_dimensions['D'].width = 100
worksheet.column_dimensions['E'].width = 120#####调整列宽
worksheet.append( ['通知来源','通知链接','通知发布时间','通知标题','通知简要'])##表头
workbook.save('./ret.xlsx')
row_n = 2##记录行号

while True:
    print("选择你要爬取的网站")
    print("1 : 山大视点-山大要闻;")
    print("2 : 本科生院-工作通知;")
    print("3 : 计算机学院-本科教育;")
    print("4 : 在下一个自动爬取点自动爬取；")
    print("5 : 停止;")
    print("6 : 重置ret.xlsx文件.")#功能选项
    choice = (int)(input())
    if choice == 4:#自动爬取
        print("程序将在上午7点，正午12点，下午6点，晚上10点自动运行.")
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"当前时间：{current_time}")
        print("Ctrl+C以停止程序.")
        scheduler = BlockingScheduler()
        try:
            scheduler.add_job(Auto_fetch,'cron',hour = 7,minute = 0)
            scheduler.add_job(Auto_fetch,'cron',hour = 12,minute =0)
            scheduler.add_job(Auto_fetch,'cron',hour = 18,minute = 0)
            scheduler.add_job(Auto_fetch,'cron',hour = 22,minute = 0)
            scheduler.start()
        except KeyboardInterrupt:#停止信号
            print("程序已终止.")
    elif choice == 5:
        print("已停止.")
        workbook = load_workbook(filename = r'./ret.xlsx')
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=1,max_row=1000,min_col=1,max_col=50):
            for cell in row:
                cell.font = ft
                cell.alignment = alignment
                cell.number_format = number_format
        workbook.save('./ret.xlsx')
        break#格式化后停止
    else  :
        if choice == 6:#重置操作
            workbook = load_workbook(filename = r'./ret.xlsx')
            worksheet = workbook.active
            del workbook['Sheet']
            workbook = op.Workbook()
            worksheet = workbook.active
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 60
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 100
            worksheet.column_dimensions['E'].width = 120#####调整列宽
            worksheet.append( ['通知来源','通知链接','通知发布时间','通知标题','通知简要'])##表头
            row_n = 2##记录行号
            for row in worksheet.iter_rows(min_row=1,max_row=1000,min_col=1,max_col=50):
                for cell in row:
                    cell.font = ft
                    cell.alignment = alignment
                    cell.number_format = number_format
            workbook.save('ret.xlsx')#格式化
        else:#爬取操作
            numbers = (int)(input("你期望的数量（以页计）："))
            workbook = load_workbook(filename = r'./ret.xlsx')
            worksheet = workbook.active
            row_n = worksheet.max_row + 1
            workbook.save('ret.xlsx')
            fetch(numbers,choice,row_n)

    

